"""
Import massivo di record Inventory da un file Excel.

Mappatura colonne Excel -> campi Inventory (verificata sul modello reale):
  Tipo         -> type      (FK core.InventoryType, lookup per "key"; creato se non esiste)
  Customer     -> customer  (FK crm.Customer, lookup per "code")
  Sito         -> site      (FK crm.Site, lookup per "display_name" dentro il customer)
  Nome         -> name
  Hostname     -> hostname
  Knumber      -> knumber
  IP Locale    -> local_ip
  IP SRSA      -> srsa_ip
  Stato        -> status    (FK core.InventoryStatus, lookup per "key": in_use/maintenance/repair)
  Seriale      -> serial_number
  Modello      -> model
  Produttore   -> manufacturer
  Nota         -> notes
  VNC Password -> vnc_pwd
  Utente OS    -> os_user
  Password OS  -> os_pwd

Deduplica: se esiste già un Inventory NON eliminato con lo stesso hostname,
serial_number o local_ip, la riga viene saltata e segnalata nel report finale.

Uso:
  python manage.py import_inventory /path/al/file.xlsx --sheet Esempio
  python manage.py import_inventory /path/al/file.xlsx --sheet Esempio --dry-run
"""

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q

from core.models import InventoryStatus, InventoryType
from crm.models import Customer, Site
from inventory.models import Inventory

VALID_STATUS_KEYS = {"in_use", "maintenance", "repair"}

COLUMN_MAP = {
    "Nome": "name",
    "Hostname": "hostname",
    "Knumber": "knumber",
    "IP Locale": "local_ip",
    "IP SRSA": "srsa_ip",
    "Seriale": "serial_number",
    "Modello": "model",
    "Produttore": "manufacturer",
    "Nota": "notes",
    "VNC Password": "vnc_pwd",
    "Utente OS": "os_user",
    "Password OS": "os_pwd",
}

DEDUPE_FIELDS = ["hostname", "serial_number", "local_ip"]


def _norm_key(value):
    """Normalizza una stringa a formato 'key' (minuscolo, underscore)."""
    if value is None:
        return None
    return str(value).strip().lower().replace(" ", "_")


class Command(BaseCommand):
    help = "Importa dispositivi da un file Excel in inventory.Inventory"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Percorso del file .xlsx")
        parser.add_argument(
            "--sheet", type=str, default="Esempio", help="Nome del foglio da importare"
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula l'import senza scrivere nulla sul DB",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File non trovato: {excel_path}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(
                f"Foglio '{sheet_name}' non trovato. Disponibili: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data_rows = rows[1:]

        created, skipped_dupe, errors = 0, [], []

        for i, row in enumerate(data_rows, start=2):  # riga 1 = header
            row_dict = dict(zip(headers, row))

            if all(v is None for v in row_dict.values()):
                continue  # riga vuota

            label = row_dict.get("Nome") or row_dict.get("Hostname") or f"riga {i}"

            # --- Customer ---
            customer_code = row_dict.get("Customer")
            try:
                customer = Customer.objects.get(code=customer_code)
            except Customer.DoesNotExist:
                errors.append((i, label, f"Customer con code='{customer_code}' non trovato"))
                continue

            # --- Site ---
            site_name = row_dict.get("Sito")
            site = None
            if site_name:
                site = Site.objects.filter(customer=customer, display_name=site_name).first()
                if site is None:
                    errors.append(
                        (i, label, f"Site '{site_name}' non trovato per il customer '{customer_code}'")
                    )
                    continue

            # --- Type (auto-creato se nuovo) ---
            tipo_raw = row_dict.get("Tipo")
            tipo_key = _norm_key(tipo_raw)
            if not tipo_key:
                errors.append((i, label, "Campo 'Tipo' mancante"))
                continue
            inv_type, type_created = InventoryType.objects.get_or_create(
                key=tipo_key, defaults={"label": str(tipo_raw).strip().title()}
            )

            # --- Status (deve essere uno dei valori validi) ---
            stato_raw = row_dict.get("Stato")
            stato_key = _norm_key(stato_raw)
            if stato_key not in VALID_STATUS_KEYS:
                errors.append(
                    (i, label, f"Stato '{stato_raw}' non valido (validi: {sorted(VALID_STATUS_KEYS)})")
                )
                continue
            inv_status = InventoryStatus.objects.get(key=stato_key)

            # --- Payload campi diretti ---
            payload = {"customer": customer, "site": site, "type": inv_type, "status": inv_status}
            for excel_col, model_field in COLUMN_MAP.items():
                payload[model_field] = row_dict.get(excel_col)

            # --- Deduplica ---
            dupe_q = None
            for f in DEDUPE_FIELDS:
                val = payload.get(f)
                if val:
                    cond = {f: val}
                    dupe_q = (dupe_q | Q(**cond)) if dupe_q is not None else Q(**cond)
            if dupe_q is not None:
                existing = Inventory.objects.filter(dupe_q, deleted_at__isnull=True).first()
                if existing is not None:
                    matched_on = [f for f in DEDUPE_FIELDS if payload.get(f) and getattr(existing, f, None) == payload.get(f)]
                    skipped_dupe.append((i, label, existing.pk, matched_on))
                    continue

            if dry_run:
                self.stdout.write(f"[DRY-RUN] riga {i}: creerei {payload}")
                created += 1
                continue

            with transaction.atomic():
                Inventory.objects.create(**payload)
            created += 1

        # --- Report finale ---
        self.stdout.write(self.style.SUCCESS(f"\nImportati: {created}"))

        if skipped_dupe:
            self.stdout.write(self.style.WARNING(f"Saltati per duplicato ({len(skipped_dupe)}):"))
            for riga, nome, existing_id, matched_on in skipped_dupe:
                self.stdout.write(
                    f"  - riga {riga} ({nome}): già presente come Inventory #{existing_id} "
                    f"(match su: {', '.join(matched_on)})"
                )

        if errors:
            self.stdout.write(self.style.ERROR(f"Errori/non importati ({len(errors)}):"))
            for riga, nome, msg in errors:
                self.stdout.write(f"  - riga {riga} ({nome}): {msg}")
