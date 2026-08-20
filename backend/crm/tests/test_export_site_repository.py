"""Test per CustomerViewSet.export_site_repository: export xlsx flat
(cliente → siti → inventory) scoped a un solo cliente.
"""
import io

import pytest

from crm.models import Customer, Site
from inventory.models import Inventory

pytestmark = pytest.mark.django_db


def _make_customer(customer_status, **kwargs):
    defaults = {"name": "ACME Hospital", "status": customer_status}
    defaults.update(kwargs)
    return Customer.objects.create(**defaults)


def test_export_returns_xlsx_content_type_and_disposition(api_client, superuser, customer_status):
    api_client.force_authenticate(user=superuser)
    customer = _make_customer(customer_status)

    resp = api_client.get(f"/api/customers/{customer.id}/export-site-repository/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp["Content-Disposition"].startswith("attachment;")
    assert resp["Content-Disposition"].endswith('.xlsx"')
    # Uno xlsx e' uno zip: signature PK\x03\x04.
    assert resp.content[:2] == b"PK"


def test_export_filename_includes_customer_code(api_client, superuser, customer_status):
    api_client.force_authenticate(user=superuser)
    customer = _make_customer(customer_status)

    resp = api_client.get(f"/api/customers/{customer.id}/export-site-repository/")
    assert customer.code in resp["Content-Disposition"]


def test_export_one_row_per_asset_with_customer_and_site_columns(
    api_client, superuser, customer_status, site_status, inventory_status, inventory_type,
):
    from openpyxl import load_workbook

    api_client.force_authenticate(user=superuser)
    customer = _make_customer(customer_status, name="ASL 2 Savonese")
    site = Site.objects.create(customer=customer, name="Ospedale Centrale", status=site_status, city="Savona")
    Inventory.objects.create(
        customer=customer, site=site, name="PACS-01", type=inventory_type, status=inventory_status,
        hostname="pacs01.local",
    )
    Inventory.objects.create(
        customer=customer, site=site, name="PACS-02", type=inventory_type, status=inventory_status,
        hostname="pacs02.local",
    )

    resp = api_client.get(f"/api/customers/{customer.id}/export-site-repository/")
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    header = [c.value for c in ws[1]]
    assert header[:5] == ["Cliente codice", "Cliente nome", "Cliente città", "Cliente provincia", "Cliente stato"]
    assert "Hostname" in header
    hostname_idx = header.index("Hostname")
    customer_name_idx = header.index("Cliente nome")
    site_name_idx = header.index("Sito nome")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2  # una riga per asset, non per sito/cliente
    hostnames = {row[hostname_idx] for row in rows}
    assert hostnames == {"pacs01.local", "pacs02.local"}
    for row in rows:
        assert row[customer_name_idx] == "ASL 2 Savonese"
        assert row[site_name_idx] == "Ospedale Centrale"


def test_export_customer_without_assets_still_produces_a_row(
    api_client, superuser, customer_status, site_status,
):
    from openpyxl import load_workbook

    api_client.force_authenticate(user=superuser)
    customer = _make_customer(customer_status, name="Cliente senza asset")
    Site.objects.create(customer=customer, name="Sede unica", status=site_status)

    resp = api_client.get(f"/api/customers/{customer.id}/export-site-repository/")
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    header = [c.value for c in ws[1]]
    assert rows[0][header.index("Sito nome")] == "Sede unica"


def test_export_excludes_secret_fields(
    api_client, superuser, customer_status, inventory_status, inventory_type,
):
    """Le credenziali (os_pwd/app_pwd/vnc_pwd) non devono mai comparire
    nell'export, indipendentemente dai permessi dell'utente: un xlsx gira
    via email molto più facilmente di una schermata in-app."""
    api_client.force_authenticate(user=superuser)
    customer = _make_customer(customer_status, name="Cliente con segreti")
    Inventory.objects.create(
        customer=customer, name="Server", type=inventory_type, status=inventory_status,
        os_pwd="supersegreto", app_pwd="anchepiusegreto", vnc_pwd="vncsegreto",
    )

    resp = api_client.get(f"/api/customers/{customer.id}/export-site-repository/")
    assert b"supersegreto" not in resp.content
    assert b"anchepiusegreto" not in resp.content
    assert b"vncsegreto" not in resp.content


def test_export_requires_authentication(api_client, customer_status):
    customer = _make_customer(customer_status)
    resp = api_client.get(f"/api/customers/{customer.id}/export-site-repository/")
    assert resp.status_code in (401, 403)
